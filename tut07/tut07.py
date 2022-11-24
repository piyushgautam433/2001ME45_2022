import os  # Import Module
os.system('cls')
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side

import numpy as np
import math
import pandas as pd
# importing required libraries and OS module

started_time = datetime.now()

mod = 5000
# mod value entered

# Enter the path of folder where Input files are stored i.e. Input Folder
path_in = r"C:\Users\HP\Documents\GitHub\2001ME45_2022\tut07\input"
path_out = r"C:\Users\HP\Documents\GitHub\2001ME45_2022\tut07\output"
# Change the directory
os.chdir(path_in)

# iterate through all file
for file in os.listdir():
    # Check whether file is in xlsx format or not
    if file.endswith(".xlsx"):
        file_path = f"{path_in}\{file}"
        try:
            input = pd.read_excel(file)
        except:
            print("Error! input file present is not in directory")
            print("Error! Close if input file is open and try again")
            exit()

        # loading input excel file
        output = pd.DataFrame(input)

        output.loc[0, 'U_Avg'] = input['U'].mean() #creating column of mean values
        output.loc[0, 'V_Avg'] = input['V'].mean()
        output.loc[0, 'W_Avg'] = input['W'].mean()

        output['U\' = U - U avg'] = input['U'] - input['U'].mean()
        output['V\' = V - V avg'] = input['V'] - input['V'].mean()
        output['W\' = W - W avg'] = input['W'] - input['W'].mean()

        def get_octant(x, y):
            if x >= 0 and y >= 0:
                return 1
            elif x < 0 and y >= 0:
                return 2
            elif x >= 0 and y < 0:
                return 4
            else:
                return 3

        # Finding Octant Values
        output['Octant'] = [(1 if row['W\' = W - W avg'] > 0 else -1) *
                            get_octant(row['U\' = U - U avg'], row['V\' = V - V avg']) for _, row in output.iterrows()]

        result = pd.concat([output], axis=1)

        # setting counters initially equals to zero
        count_p1 = 0
        count_p2 = 0
        count_p3 = 0
        count_p4 = 0
        count_n1 = 0
        count_n2 = 0
        count_n3 = 0
        count_n4 = 0

        # Finding total number of octant values
        length = len(output['Octant'])

        for i in range(length):    
            if output.loc[i, 'Octant'] == 1:
                count_p1 = count_p1 + 1

            elif output.loc[i, 'Octant'] == 2:
                count_p2 = count_p2 + 1

            elif output.loc[i, 'Octant'] == 3:
                count_p3 = count_p3 + 1

            elif output.loc[i, 'Octant'] == 4:
                count_p4 = count_p4 + 1

            elif output.loc[i, 'Octant'] == -1:
                count_n1 = count_n1 + 1

            elif output.loc[i, 'Octant'] == -2:
                count_n2 = count_n2 + 1

            elif output.loc[i, 'Octant'] == -3:
                count_n3 = count_n3 + 1

            elif output.loc[i, 'Octant'] == -4:
                count_n4 = count_n4 + 1

        n = int(math.ceil(length / mod))

        result.loc[0, '                    '] = ""
        result.loc[1, '                     '] = "mod 5000"
        result.loc[0, 'Octant ID'] = "Overall Count"

        #counting values of +1 - -4
        result.loc[0, '+1'] = count_p1
        result.loc[0, '-1'] = count_n1
        result.loc[0, '+2'] = count_p2
        result.loc[0, '-2'] = count_n2
        result.loc[0, '+3'] = count_p3
        result.loc[0, '-3'] = count_n3
        result.loc[0, '+4'] = count_p4
        result.loc[0, '-4'] = count_n4

        for step in range(n):
            s = (step * mod)
            e = min((step + 1)*mod - 1, length - 1)
            result.loc[step+1, 'Octant ID'] = str(s) + '-' + str(e)

            result.loc[step+1, '+1'] = list(result.loc[s:e, 'Octant']).count(1)
            result.loc[step+1,
                       '-1'] = list(result.loc[s:e, 'Octant']).count(-1)
            result.loc[step+1, '+2'] = list(result.loc[s:e, 'Octant']).count(2)
            result.loc[step+1,
                       '-2'] = list(result.loc[s:e, 'Octant']).count(-2)
            result.loc[step+1, '+3'] = list(result.loc[s:e, 'Octant']).count(3)
            result.loc[step+1,
                       '-3'] = list(result.loc[s:e, 'Octant']).count(-3)
            result.loc[step+1, '+4'] = list(result.loc[s:e, 'Octant']).count(4)
            result.loc[step+1,
                       '-4'] = list(result.loc[s:e, 'Octant']).count(-4)

        result.loc['Rank 1'] = ""
        result.loc['Rank -1'] = ""
        result.loc['Rank 2'] = ""
        result.loc['Rank -2'] = ""
        result.loc['Rank 3'] = ""
        result.loc['Rank -3'] = ""
        result.loc['Rank 4'] = ""
        result.loc['Rank -4'] = ""
        result.loc['Rank1 Octant ID'] = ""
        result.loc['Rank1 Octant Name'] = ""

        List3 = []

        def Ranking(count):
            if count == 0:
                count_1 = 0
            else:
                count_1 = count - 1
            List = []

            List.append(result.loc[count_1, '+1']) #appending counts
            List.append(result.loc[count_1, '-1'])
            List.append(result.loc[count_1, '+2'])
            List.append(result.loc[count_1, '-2'])
            List.append(result.loc[count_1, '+3'])
            List.append(result.loc[count_1, '-3'])
            List.append(result.loc[count_1, '+4'])
            List.append(result.loc[count_1, '-4'])

            List1 = List.copy()

            List1.sort(reverse=True)

            List2 = []
            for i in range(0, 8):
                List2.append(List1.index(List[i], 0, 8))
            if count == 0:
                count_1 = 0
            else:
                count_1 = count - 1

            result.loc[count_1, 'Rank 1'] = List2[0] + 1 #appending ranks
            result.loc[count_1, 'Rank -1'] = List2[1] + 1
            result.loc[count_1, 'Rank 2'] = List2[2] + 1
            result.loc[count_1, 'Rank -2'] = List2[3] + 1
            result.loc[count_1, 'Rank 3'] = List2[4] + 1
            result.loc[count_1, 'Rank -3'] = List2[5] + 1
            result.loc[count_1, 'Rank 4'] = List2[6] + 1
            result.loc[count_1, 'Rank -4'] = List2[7] + 1

            Rank1 = List2.index(0, 0, 8) + 1

            List3.append(Rank1)

            if Rank1 == 1:
                result.loc[count_1, 'Rank1 Octant ID'] = "1"
                result.loc[count_1,
                           'Rank1 Octant Name'] = 'Internal outward interaction'

            elif Rank1 == 2:
                result.loc[count_1, 'Rank1 Octant ID'] = "-1"
                result.loc[count_1,
                           'Rank1 Octant Name'] = 'External outward interaction'

            elif Rank1 == 3:
                result.loc[count_1, 'Rank1 Octant ID'] = "2"
                result.loc[count_1, 'Rank1 Octant Name'] = 'External Ejection'

            elif Rank1 == 4:
                result.loc[count_1, 'Rank1 Octant ID'] = "-2"
                result.loc[count_1, 'Rank1 Octant Name'] = 'Internal Ejection'

            elif Rank1 == 5:
                result.loc[count_1, 'Rank1 Octant ID'] = "3"
                result.loc[count_1,
                           'Rank1 Octant Name'] = 'External inward interaction'

            elif Rank1 == 6:
                result.loc[count_1, 'Rank1 Octant ID'] = "-3"
                result.loc[count_1,
                           'Rank1 Octant Name'] = 'Internal inward interaction'

            elif Rank1 == 7:
                result.loc[count_1, 'Rank1 Octant ID'] = "4"
                result.loc[count_1, 'Rank1 Octant Name'] = 'Internal sweep'

            elif Rank1 == 8:
                result.loc[count_1, 'Rank1 Octant ID'] = "-4"
                result.loc[count_1, 'Rank1 Octant Name'] = 'External sweep'

        Ranking(0)

        for i in range(2, n+2):
            Ranking(i)

        result.loc[n+5, 'Rank 4'] = 'Octant ID'
        result.loc[n+5, 'Rank -4'] = 'Octant ID'
        result.loc[n+5, 'Rank1 Octant ID'] = 'Count of Rank 1 Mod Values'

        result.loc[n+6, 'Rank 4'] = '+1'
        result.loc[n+7, 'Rank 4'] = '-1'
        result.loc[n+8, 'Rank 4'] = '+2'
        result.loc[n+9, 'Rank 4'] = '-2'
        result.loc[n+10, 'Rank 4'] = '+3'
        result.loc[n+11, 'Rank 4'] = '-3'
        result.loc[n+12, 'Rank 4'] = '+4'
        result.loc[n+13, 'Rank 4'] = '-4'

        result.loc[n+6, 'Rank -4'] = 'Internal outward interaction'
        result.loc[n+7, 'Rank -4'] = 'External outward interaction'
        result.loc[n+8, 'Rank -4'] = 'External Ejection'
        result.loc[n+9, 'Rank -4'] = 'Internal Ejection'
        result.loc[n+10, 'Rank -4'] = 'External inward interaction'
        result.loc[n+11, 'Rank -4'] = 'Internal inward interaction'
        result.loc[n+12, 'Rank -4'] = 'Internal sweep'
        result.loc[n+13, 'Rank -4'] = 'External sweep'
        List4 = List3.copy()
        List3.pop(0)

        result.loc[n+6, 'Rank1 Octant ID'] = List3.count(1)
        result.loc[n+7, 'Rank1 Octant ID'] = List3.count(2)
        result.loc[n+8, 'Rank1 Octant ID'] = List3.count(3)
        result.loc[n+9, 'Rank1 Octant ID'] = List3.count(4)
        result.loc[n+10, 'Rank1 Octant ID'] = List3.count(5)
        result.loc[n+11, 'Rank1 Octant ID'] = List3.count(6)
        result.loc[n+12, 'Rank1 Octant ID'] = List3.count(7)
        result.loc[n+13, 'Rank1 Octant ID'] = List3.count(8)

        result.loc[0, '  '] = ""
        result.loc[0, '   '] = ""
        result.loc[0, '    '] = ""
        result.loc[0, '     '] = ""
        result.loc[0, '      '] = ""
        result.loc[0, '       '] = ""
        result.loc[0, '        '] = ""
        result.loc[0, '         '] = ""
        result.loc[0, '          '] = ""
        result.loc[0, '           '] = ""
        result.loc[0, '            '] = ""

        # Defining a function to print the transition for a range as specified
        def modulus(step_1, s, e):

            rows, cols = (8, 8)
            arr = [[0 for i in range(cols)] for j in range(rows)]
            s_val = str(s)
            n_2 = int(math.ceil(length / mod))
            if step_1 == n_2-1:
                e_val = str(e)
            else:
                e_val = str(e-1)
            length_1 = len(output['Octant'])
            # findig the transition values and saving it in an array
            for i in range(s, e):

                if output.loc[i, 'Octant'] == 1:

                    if output.loc[i+1, 'Octant'] == 1:
                        arr[0][0] = arr[0][0] + 1

                    elif output.loc[i+1, 'Octant'] == -1:
                        arr[0][1] = arr[0][1] + 1

                    elif output.loc[i+1, 'Octant'] == 2:
                        arr[0][2] = arr[0][2] + 1

                    elif output.loc[i+1, 'Octant'] == -2:
                        arr[0][3] = arr[0][3] + 1

                    elif output.loc[i+1, 'Octant'] == 3:
                        arr[0][4] = arr[0][4] + 1

                    elif output.loc[i+1, 'Octant'] == -3:
                        arr[0][5] = arr[0][5] + 1

                    elif output.loc[i+1, 'Octant'] == 4:
                        arr[0][6] = arr[0][6] + 1

                    elif output.loc[i+1, 'Octant'] == -4:
                        arr[0][7] = arr[0][7] + 1

                elif output.loc[i, 'Octant'] == -1:

                    if output.loc[i+1, 'Octant'] == 1:
                        arr[1][0] = arr[1][0] + 1

                    elif output.loc[i+1, 'Octant'] == -1:
                        arr[1][1] = arr[1][1] + 1

                    elif output.loc[i+1, 'Octant'] == 2:
                        arr[1][2] = arr[1][2] + 1

                    elif output.loc[i+1, 'Octant'] == -2:
                        arr[1][3] = arr[1][3] + 1

                    elif output.loc[i+1, 'Octant'] == 3:
                        arr[1][4] = arr[1][4] + 1

                    elif output.loc[i+1, 'Octant'] == -3:
                        arr[1][5] = arr[1][5] + 1

                    elif output.loc[i+1, 'Octant'] == 4:
                        arr[1][6] = arr[1][6] + 1

                    elif output.loc[i+1, 'Octant'] == -4:
                        arr[1][7] = arr[1][7] + 1

                elif output.loc[i, 'Octant'] == 2:

                    if output.loc[i+1, 'Octant'] == 1:
                        arr[2][0] = arr[2][0] + 1

                    elif output.loc[i+1, 'Octant'] == -1:
                        arr[2][1] = arr[2][1] + 1

                    elif output.loc[i+1, 'Octant'] == 2:
                        arr[2][2] = arr[2][2] + 1

                    elif output.loc[i+1, 'Octant'] == -2:
                        arr[2][3] = arr[2][3] + 1

                    elif output.loc[i+1, 'Octant'] == 3:
                        arr[2][4] = arr[2][4] + 1

                    elif output.loc[i+1, 'Octant'] == -3:
                        arr[2][5] = arr[2][5] + 1

                    elif output.loc[i+1, 'Octant'] == 4:
                        arr[2][6] = arr[2][6] + 1

                    elif output.loc[i+1, 'Octant'] == -4:
                        arr[2][7] = arr[2][7] + 1

                elif output.loc[i, 'Octant'] == -2:

                    if output.loc[i+1, 'Octant'] == 1:
                        arr[3][0] = arr[3][0] + 1

                    elif output.loc[i+1, 'Octant'] == -1:
                        arr[3][1] = arr[3][1] + 1

                    elif output.loc[i+1, 'Octant'] == 2:
                        arr[3][2] = arr[3][2] + 1

                    elif output.loc[i+1, 'Octant'] == -2:
                        arr[3][3] = arr[3][3] + 1

                    elif output.loc[i+1, 'Octant'] == 3:
                        arr[3][4] = arr[3][4] + 1

                    elif output.loc[i+1, 'Octant'] == -3:
                        arr[3][5] = arr[3][5] + 1

                    elif output.loc[i+1, 'Octant'] == 4:
                        arr[3][6] = arr[3][6] + 1

                    elif output.loc[i+1, 'Octant'] == -4:
                        arr[3][7] = arr[3][7] + 1

                elif output.loc[i, 'Octant'] == 3:

                    if output.loc[i+1, 'Octant'] == 1:
                        arr[4][0] = arr[4][0] + 1

                    elif output.loc[i+1, 'Octant'] == -1:
                        arr[4][1] = arr[4][1] + 1

                    elif output.loc[i+1, 'Octant'] == 2:
                        arr[4][2] = arr[4][2] + 1

                    elif output.loc[i+1, 'Octant'] == -2:
                        arr[4][3] = arr[4][3] + 1

                    elif output.loc[i+1, 'Octant'] == 3:
                        arr[4][4] = arr[4][4] + 1

                    elif output.loc[i+1, 'Octant'] == -3:
                        arr[4][5] = arr[4][5] + 1

                    elif output.loc[i+1, 'Octant'] == 4:
                        arr[4][6] = arr[4][6] + 1

                    elif output.loc[i+1, 'Octant'] == -4:
                        arr[4][7] = arr[4][7] + 1

                elif output.loc[i, 'Octant'] == -3:

                    if output.loc[i+1, 'Octant'] == 1:
                        arr[5][0] = arr[5][0] + 1

                    elif output.loc[i+1, 'Octant'] == -1:
                        arr[5][1] = arr[5][1] + 1

                    elif output.loc[i+1, 'Octant'] == 2:
                        arr[5][2] = arr[5][2] + 1

                    elif output.loc[i+1, 'Octant'] == -2:
                        arr[5][3] = arr[5][3] + 1

                    elif output.loc[i+1, 'Octant'] == 3:
                        arr[5][4] = arr[5][4] + 1

                    elif output.loc[i+1, 'Octant'] == -3:
                        arr[5][5] = arr[5][5] + 1

                    elif output.loc[i+1, 'Octant'] == 4:
                        arr[5][6] = arr[5][6] + 1

                    elif output.loc[i+1, 'Octant'] == -4:
                        arr[5][7] = arr[5][7] + 1

                elif output.loc[i, 'Octant'] == 4:

                    if output.loc[i+1, 'Octant'] == 1:
                        arr[6][0] = arr[6][0] + 1

                    elif output.loc[i+1, 'Octant'] == -1:
                        arr[6][1] = arr[6][1] + 1

                    elif output.loc[i+1, 'Octant'] == 2:
                        arr[6][2] = arr[6][2] + 1

                    elif output.loc[i+1, 'Octant'] == -2:
                        arr[6][3] = arr[6][3] + 1

                    elif output.loc[i+1, 'Octant'] == 3:
                        arr[6][4] = arr[6][4] + 1

                    elif output.loc[i+1, 'Octant'] == -3:
                        arr[6][5] = arr[6][5] + 1

                    elif output.loc[i+1, 'Octant'] == 4:
                        arr[6][6] = arr[6][6] + 1

                    elif output.loc[i+1, 'Octant'] == -4:
                        arr[6][7] = arr[6][7] + 1

                elif output.loc[i, 'Octant'] == -4:

                    if output.loc[i+1, 'Octant'] == 1:
                        arr[7][0] = arr[7][0] + 1

                    elif output.loc[i+1, 'Octant'] == -1:
                        arr[7][1] = arr[7][1] + 1

                    elif output.loc[i+1, 'Octant'] == 2:
                        arr[7][2] = arr[7][2] + 1

                    elif output.loc[i+1, 'Octant'] == -2:
                        arr[7][3] = arr[7][3] + 1

                    elif output.loc[i+1, 'Octant'] == 3:
                        arr[7][4] = arr[7][4] + 1

                    elif output.loc[i+1, 'Octant'] == -3:
                        arr[7][5] = arr[7][5] + 1

                    elif output.loc[i+1, 'Octant'] == 4:
                        arr[7][6] = arr[7][6] + 1

                    elif output.loc[i+1, 'Octant'] == -4:
                        arr[7][7] = arr[7][7] + 1

            n = int(math.ceil(length / mod))

            result.loc[n+6 + ((step_1 + 1)*13) - 8, '   '] = "From"
            # If step_1 == -1 then overall transition table will print
            if step_1 == -1:
                result.loc[n+4 + ((step_1 + 1)*13) - 8,
                           '    '] = "Overall Transition Count"
                result.loc[n+5 + ((step_1 + 1)*13) - 8,
                           '     '] = "To"
            else:
                # if step_1 is >= 0 then the transition values as per step_1 range would be printed
                result.loc[n+4 + ((step_1 + 1)*13) - 8,
                           '    '] = "Mod Transition Count"
                result.loc[n+5 + ((step_1 + 1)*13) - 8,
                           '    '] = s_val + "-" + e_val
                result.loc[n+5 + ((step_1 + 1)*13) - 8,
                           '     '] = "To"
            # Specifying an 9*9 matrix in which 8*8 is our data and 1st column and row are fixed index of the table
            rows_1, cols_1 = (9, 9)
            arr_1 = [[0 for i in range(cols_1)] for j in range(rows_1)]
            arr_1[0][0] = "count"
            arr_1[1][0] = "+1"
            arr_1[2][0] = "-1"
            arr_1[3][0] = "+2"
            arr_1[4][0] = "-2"
            arr_1[5][0] = "+3"
            arr_1[6][0] = "-3"
            arr_1[7][0] = "+4"
            arr_1[8][0] = "-4"
            arr_1[0][1] = "+1"
            arr_1[0][2] = "-1"
            arr_1[0][3] = "+2"
            arr_1[0][4] = "-2"
            arr_1[0][5] = "+3"
            arr_1[0][6] = "-3"
            arr_1[0][7] = "+4"
            arr_1[0][8] = "-4"

            counter_1 = 0
            counter_2 = 0
            while counter_1 < 8:
                while counter_2 < 8:
                    arr_1[counter_1 + 1][counter_2 +
                                         1] = arr[counter_1][counter_2]
                    counter_2 = counter_2 + 1
                counter_2 = 0
                counter_1 = counter_1 + 1

            counter_3 = 0
            counter_4 = 0
            # logic to fill up the table in output file
            while counter_3 < 9:

                result.loc[n + 6 + counter_3 + ((step_1 + 1)*13) - 8,
                           '    '] = arr_1[counter_3][counter_4]
                result.loc[n + 6 + counter_3 +
                           ((step_1 + 1)*13) - 8, '     '] = arr_1[counter_3][counter_4 + 1]
                result.loc[n + 6 + counter_3 +
                           ((step_1 + 1)*13) - 8, '      '] = arr_1[counter_3][counter_4 + 2]
                result.loc[n + 6 + counter_3 +
                           ((step_1 + 1)*13) - 8, '       '] = arr_1[counter_3][counter_4 + 3]
                result.loc[n + 6 + counter_3 +
                           ((step_1 + 1)*13) - 8, '        '] = arr_1[counter_3][counter_4 + 4]
                result.loc[n + 6 + counter_3 +
                           ((step_1 + 1)*13) - 8, '         '] = arr_1[counter_3][counter_4 + 5]
                result.loc[n + 6 + counter_3 +
                           ((step_1 + 1)*13) - 8, '          '] = arr_1[counter_3][counter_4 + 6]
                result.loc[n + 6 + counter_3 +
                           ((step_1 + 1)*13) - 8, '           '] = arr_1[counter_3][counter_4 + 7]
                result.loc[n + 6 + counter_3 +
                           ((step_1 + 1)*13) - 8, '            '] = arr_1[counter_3][counter_4 + 8]
                counter_3 = 1 + counter_3

        last_value = length
        modulus(-1, 0, last_value-1)

        n_1 = int(math.ceil(length / mod))

        for step_1 in range(n_1):
            s = (step_1 * mod)
            e = min((step_1 + 1)*mod, length-1)
            modulus(step_1, s, e)

        # specifying few column index's as per requirement
        result["                         "] = ""
        result["Octant #"] = ""
        result["Longest Subsequence Length"] = ""
        result["Count"] = ""

        result["                          "] = ""
        result["Octant ##"] = ""
        result["Longest Subsequence Length "] = ""
        result["Count "] = ""

        # Filling the Count column as required in the output file
        result.loc[0, "Octant #"] = "+1"
        result.loc[1, "Octant #"] = "-1"
        result.loc[2, "Octant #"] = "+2"
        result.loc[3, "Octant #"] = "-2"
        result.loc[4, "Octant #"] = "+3"
        result.loc[5, "Octant #"] = "-3"
        result.loc[6, "Octant #"] = "+4"
        result.loc[7, "Octant #"] = "-4"

        #result = pd.concat([output], axis=1)
        # finding the total no. of octant values
        length = len(output['Octant'])
        # list = [""]*length
        # for i in range(length):
        #     list[i] = output.loc[i, 'Octant']

        # Declaring a empty string
        string1 = ""
        # Saving the octant values in the string
        for i in range(length):
            if output.loc[i, 'Octant'] > 0:
                a = str(output.loc[i, 'Octant'])
                string1 = string1 + a
            if output.loc[i, 'Octant'] < 0:
                # if the string is 1-222-33
                # If we have to count no. consecutive occurences of "2"
                # the occurding to string it will be displayed as 3, but answer is 2
                # as the "-" sign of 1st "=2" will not be considered in string
                # So tho make it unique, writing the string as : 1*-2-22*-3*3
                # i.e. "*"" + "negative number" + "*""
                a = "*" + str(output.loc[i, 'Octant']) + "*"
                string1 = string1 + a
        string_1 = ""
        for i in range(length):
            if output.loc[i, 'Octant'] > 0:
                a = str(output.loc[i, 'Octant'])
                string_1 = string_1 + a
            elif output.loc[i, 'Octant'] == -1:
                a = "5"
                string_1 = string_1 + a
            elif output.loc[i, 'Octant'] == -2:
                a = "6"
                string_1 = string_1 + a
            elif output.loc[i, 'Octant'] == -3:
                a = "7"
                string_1 = string_1 + a
            elif output.loc[i, 'Octant'] == -4:
                a = "8"
                string_1 = string_1 + a

        # defing function to find maximum consecutive repetition of string2 in string1

        def maxRepeating(string1, string2):

            # Stores the count of consecutive
            # occurrences of string2 in string1
            Count1 = string1.count(string2)

            # Concatenate string2 Count1 times
            Count2 = string2 * Count1

            # Iterate over the string string1
            # while Count2 is not present in string1
            while (Count2 not in string1):

                # Update Count1
                Count1 -= 1

            # Update Count2
                Count2 = string2 * Count1

            return Count1

        # Defining a function to find maximum occurences of a string in another string

        def occurrences(string1, string2):

            # Stores the count of consecutive
            # occurrences of string2 in string1
            Count_01 = string1.count(string2)
            return Count_01

        # Driver Code
        # Finding maximum no. of consecutive occurences of 1 in octant value string
        if __name__ == "__main__":
            string = "1"
            result.loc[0, "Longest Subsequence Length"] = maxRepeating(
                string1, string)
        # Finding the total no. of occurences of longest string of 1 in Octant value string
        string_01 = "1" * result.loc[0, "Longest Subsequence Length"]
        result.loc[0, "Count"] = occurrences(string1, string_01)

        # Finding maximum no. of consecutive occurences of *-1* in octant value string
        if __name__ == "__main__":
            string2 = "*-1*"
            result.loc[1, "Longest Subsequence Length"] = maxRepeating(
                string1, string2)
        # Finding the total no. of occurences of longest string of *-1* in Octant value string
        string_02 = "*-1*" * result.loc[1, "Longest Subsequence Length"]
        result.loc[1, "Count"] = occurrences(string1, string_02)

        if __name__ == "__main__":
            str3 = "2"
            result.loc[2, "Longest Subsequence Length"] = maxRepeating(
                string1, str3)

        string_03 = "2" * result.loc[2, "Longest Subsequence Length"]
        result.loc[2, "Count"] = occurrences(string1, string_03)

        if __name__ == "__main__":
            string2 = "*-2*"
            result.loc[3, "Longest Subsequence Length"] = maxRepeating(
                string1, string2)

        string_04 = "*-2*" * result.loc[3, "Longest Subsequence Length"]
        result.loc[3, "Count"] = occurrences(string1, string_04)

        if __name__ == "__main__":
            string2 = "3"
            result.loc[4, "Longest Subsequence Length"] = maxRepeating(
                string1, string2)

        string_05 = "3" * result.loc[4, "Longest Subsequence Length"]
        result.loc[4, "Count"] = occurrences(string1, string_05)

        if __name__ == "__main__":
            string2 = "*-3*"
            result.loc[5, "Longest Subsequence Length"] = maxRepeating(
                string1, string2)

        string_06 = "*-3*" * result.loc[5, "Longest Subsequence Length"]
        result.loc[5, "Count"] = occurrences(string1, string_05)

        if __name__ == "__main__":
            string2 = "4"
            result.loc[6, "Longest Subsequence Length"] = maxRepeating(
                string1, string2)

        string_06 = "4" * result.loc[6, "Longest Subsequence Length"]
        result.loc[6, "Count"] = occurrences(string1, string_06)

        if __name__ == "__main__":
            string2 = "*-4*"
            result.loc[7, "Longest Subsequence Length"] = maxRepeating(
                string1, string2)

        string_07 = "*-4*" * result.loc[7, "Longest Subsequence Length"]
        result.loc[7, "Count"] = occurrences(string1, string_07)

        result.loc["Output ##"] = ""
        result.loc["Longest Subsequence Length "] = ""
        result.loc["Count "] = ""

        def Longest_Subsequence_Length(i, j):
            result.loc[j, "Octant ##"] = result.loc[i, "Octant #"]
            result.loc[j, "Longest Subsequence Length "] = result.loc[i,
                                                                      "Longest Subsequence Length"]
            result.loc[j, "Count "] = result.loc[i, "Count"]

            result.loc[j+1, "Octant ##"] = "Time"
            result.loc[j+1, "Longest Subsequence Length "] = "From"
            result.loc[j+1, "Count "] = "To"
            length_1 = len(string_1)
            if i == 0:
                _A_ = "1" * result.loc[i, "Longest Subsequence Length"]
            elif i == 1:
                _A_ = "5" * result.loc[i, "Longest Subsequence Length"]
            elif i == 2:
                _A_ = "2" * result.loc[i, "Longest Subsequence Length"]
            elif i == 3:
                _A_ = "6" * result.loc[i, "Longest Subsequence Length"]
            elif i == 4:
                _A_ = "3" * result.loc[i, "Longest Subsequence Length"]
            elif i == 5:
                _A_ = "7" * result.loc[i, "Longest Subsequence Length"]
            elif i == 6:
                _A_ = "4" * result.loc[i, "Longest Subsequence Length"]
            elif i == 7:
                _A_ = "8" * result.loc[i, "Longest Subsequence Length"]
            Start = 0
            Count_0 = j
            x1 = 0
            var = 0
            while var < result.loc[i, "Count"]:

                x = string_1.find(_A_, Start, length_1)

                x1 = x + result.loc[i, "Longest Subsequence Length"]
                result.loc[Count_0 + 2,
                           "Longest Subsequence Length "] = result.loc[x, "T"]
                result.loc[Count_0 + 2, "Count "] = result.loc[x1 - 1, "T"]
                Count_0 += 1
                Start = x + 1
                var += 1

        Longest_Subsequence_Length(0, 0)
        try:
            for var_1 in range(1, 8):
                j_1 = 0
                z = 0
                k = 0
                while z < var_1:

                    j_1 += result.loc[z, "Count"]
                    z += 1
                    k += 1
                if k > 0:
                    j_1 += (k)*2
                
                Longest_Subsequence_Length(var_1, j_1)
                var_1 += 1
        except:
            print("some error")

        j_2 = 0
        j_3 = 8
        for i in range(j_3):
            j_2 += result.loc[i, "Count"]

        file_1 = file[:-5]
        print(file_1)
        mod_1 = str(mod)
        # Writing a pandas DataFrame to excel file
        try:
            result.to_excel(path_out + '/' + file_1 +
                            '_vel_octant_analysis_mod_' + mod_1 + '.xlsx')
        except:
            print("Error")

        os.chdir(path_out)
        wb = openpyxl.load_workbook(
            file_1 + '_vel_octant_analysis_mod_' + mod_1 + '.xlsx')  # path to the Excel file
        ws = wb['Sheet1']
        fill_cell = PatternFill(patternType='solid',
                                fgColor='FFF000')  # You can give the hex code for different color

        n1 = len(List4)
        for i in range(n1):
            N1 = str(openpyxl.utils.cell.get_column_letter(23+List4[i]))
            N2 = str(i+2)
            N3 = N1 + N2
            ws[N3].fill = fill_cell

        def cell_border(ws, cell_range):
            rows = ws[cell_range]
            side = Side(border_style='thin', color="FF000000")
            rows = list(rows)
            for y_pos, cells in enumerate(rows):
                for x_pos, cell in enumerate(cells):
                    border = Border(left=cell.border.left, right=cell.border.right,
                                    top=cell.border.top, bottom=cell.border.bottom)
                    border.left = side
                    border.right = side
                    border.top = side
                    border.bottom = side
                    cell.border = border
        N4 = "AG" + str(n+2)
        cell_border(ws, "O1:" + N4)
        cell_border(ws, "AD11:AF19")
        cell_border(ws, "AJ4:AR12")
        cell_border(ws, "AT1:AV9")

        Alpha = ["AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR"]

        for i in range(0, 8, 1):
            lis = []
            for j in range(0, 8, 1):
                lis.append(result[" "*(j+5)][i+3])

            maxm = max(lis)
            for j in range(0, 8, 1):
                if (result[" "*(j+5)][i+3] == maxm):
                    ws[Alpha[j] + str(i+5)].fill = fill_cell
        n_2 = int(math.ceil(length / mod))

        for k in range(0, n_2, 1):
            for i in range(0, 8, 1):
                lis2 = []
                for j in range(0, 8, 1):
                    lis2.append(result[" "*(j+5)][16+(13*k) + i])

                maxm = max(lis2)
                for j in range(0, 8, 1):
                    if (result[" "*(j+5)][16+(13*k) + i] == maxm):
                        ws[Alpha[j] + str(16+(13*k) + i+2)].fill = fill_cell

        for i in range(n_1):
            N5 = "AJ" + str(17 + i*13)
            N6 = ":AR" + str(25 + i*13)
            cell_border(ws, N5 + N6)
        N7 = "AX1"
        N8 = ":AZ" + str(j_2 + 17)
        cell_border(ws, N7 + N8)

        wb.save(file_1 + '_vel_octant_analysis_mod_' + mod_1 + '.xlsx')
        os.chdir(path_in)


end_Time = datetime.now()
print('Duration of Program Execution: {}'.format(end_Time - started_time))
