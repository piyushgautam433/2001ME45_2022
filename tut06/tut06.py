import csv #importing important libraries

import pandas as pd

from datetime import datetime #import calender
import openpyxl
starting_time = datetime.now()

def attendance_report(): #function 
    try:
        inp_file = pd.read_csv('input_attendance.csv') #attendence file important
        inp = inp_file.fillna("Random") #Random in empty rows
    except:
        print("File not found")  
    
    
    try:
        rollno_inp=pd.read_csv('input_registered_students.csv')  #registered students file
    except:
        print('File containing name of all students is missing!')

     # mr = sheet_input.max_row
    mc=sum(1 for row in open("input_registered_students.csv"))  #mc variable is sum in registered studets
    mc_consolidated=sum(1 for row in open("input_attendance.csv")) #for input attendence csv
    total_dates=list()
    for j in range(0,mc_consolidated-1): #split year, month, day
        day=inp.at[j,'Timestamp'].split()[0].split('-')[0]
        month=inp.at[j,'Timestamp'].split()[0].split('-')[1]
        year=inp.at[j,'Timestamp'].split()[0].split('-')[2]
        date=datetime.strptime(f'{year}-{month}-{day}', "%Y-%m-%d").date()  #finding day according to day
        day_name=date.strftime("%A")
        if day_name=="Monday" or day_name=="Thursday": 
            if inp.at[j,'Timestamp'].split()[0] not in total_dates: #considering monday and thrusday only in total dates
                total_dates.append(inp.at[j,'Timestamp'].split()[0])
    # max_att=0
    # max_roll=""
    fileName_consolidated=".\output\\attendance_report_consolidated.xlsx" #putting in output file

    output_file=openpyxl.Workbook() #output file
    output=output_file.active
    output_file.save(fileName_consolidated) #saving in consolidated file

    attend_consolidated=pd.read_excel(fileName_consolidated)
    for i in range(0,mc-1):
        present=0 #totsl present
        date_index=1 #data index
        rollno=rollno_inp.at[i,'Roll No']  #put roll no. in roll no section
        fileName=".\output\\"+rollno+'.xlsx' #in this file output should have
        output_file=openpyxl.Workbook() #open workbook excel
        output=output_file.active 
        output_file.save(fileName)  #saving file name
        out=pd.read_excel(fileName)  #read file
        for main_dates in total_dates: 
            duplicated_attendance=0 #whether attendence is duplicate
            t_lec,t_lec_act,t_lec_fake,t_lec_abs,percent=len(total_dates),0,0,0,0
            t_lec_count=0
            for j in range(0,mc_consolidated-1):  #for loop
                if inp.at[j,'Attendance'].split()[0]==rollno:  #if roll no. 
                    if inp.at[j,'Timestamp'].split()[0] == main_dates: #if lies in main dates
                        t_lec_count+=1  # add 1 in total lectures
                        time=inp.at[j,'Timestamp'].split()[1]
                        hour=time.split(':')[0]
                        minutes=time.split(':')[1]
                        if ((hour=='14') or (hour=='15' and minutes=='00')): #present should be there in these time for real attendence count
                            if t_lec_act==0: 
                                t_lec_act+=1
                            else:
                                duplicated_attendance+=1 #if not real than duplicate
                        else:
                            t_lec_fake+=1 #else fake attendence
            
            out.at[0,'Roll']=rollno
            out.at[0,'Name']=rollno_inp.at[i,'Name']
            out.at[date_index,'Dates']=main_dates
            attend_consolidated.at[i+1,'Roll']=rollno
            attend_consolidated.at[i+1,'Name']=rollno_inp.at[i,'Name']
            out.at[date_index,'Total Attendance Count']=t_lec_count
            attend_consolidated.at[i+1,f'{main_dates}']='P' if t_lec_act>0 else 'A'
            if t_lec_act>0: present+=1 
            out.at[date_index,'Real']=t_lec_act  #real lectures. they will be tjhe only ones which are counted
            out.at[date_index,'Duplicate']=duplicated_attendance #duplicate attendence
            out.at[date_index,'Invalid']= t_lec_fake #fake
            out.at[date_index,'Absent']=1 if t_lec_act==0 else 0 #absent column 
            date_index+=1
        attend_consolidated.at[i+1,'Actual Lecture Taken']=len(total_dates)  #total lectures taken 
        attend_consolidated.at[i+1,'Total Real']=present  #toal real
        attend_consolidated.at[i+1,'% Attendance']=round(present/len(total_dates)*100,2) #percent attendence
        out.to_excel(fileName,index=False)
        if i==int(0.2*mc): print("20% /complete")  #20% compiling done
        if i==int(0.4*mc): print("40% /complete")  #40% compiling done
        if i==int(0.6*mc): print("60% /complete")  #60% compiling done
        if i==int(0.8*mc): print("80% /complete")  #80% compiling done
        if i==int(mc-2): print("100% /complete")   #100% compiling done

    attend_consolidated.to_excel(fileName_consolidated,index=False)
# ,,,,,, (attendance_count_actual/total_lecture_taken) 2 digit decimal
